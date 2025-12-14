# in assignments/services.py

from collections import defaultdict
from students.models import Student
from subjects.services import get_subject_code, get_subject_display_name
from subjects.models import PraktikumType
from praktikums_lehrkraft.models import PraktikumsLehrkraft
import re
from schools.services import get_reachable_schools


def aggregate_demand():
    """
    Calculates the total forecasted demand for the upcoming academic year.
    """
    demand_counts = defaultdict(int)
    code_to_display_map = {"N/A": "N/A"}

    unplaced_students = Student.objects.filter(
        placement_status="UNPLACED"
    ).select_related("primary_subject", "didactic_subject_3")

    for student in unplaced_students:
        process_student_demand(student, demand_counts, code_to_display_map)

    return format_demand(demand_counts, code_to_display_map)


def process_student_demand(student, demand_counts, code_to_display_map):
    """Processes a single student and updates demand counts and display mapping."""
    program_type = student.program

    # Get the specific subject for each practicum type
    sfp_subject_name = (
        student.primary_subject.name if student.primary_subject else "N/A"
    )
    zsp_subject_name = (
        student.didactic_subject_3.name if student.didactic_subject_3 else "N/A"
    )

    # --- Rule 1: PDP I ---
    if student.pdp1_completed_date is None:
        key = "PDP_I", program_type, "N/A"
        demand_counts[key] += 1

    # --- Rule 2: PDP II ---
    if student.pdp1_completed_date is not None and student.pdp2_completed_date is None:
        key = "PDP_II", program_type, "N/A"
        demand_counts[key] += 1

    # --- Rule 3: SFP ---
    if student.sfp_completed_date is None and sfp_subject_name != "N/A":
        add_practicum_demand(
            "SFP", program_type, sfp_subject_name, demand_counts, code_to_display_map
        )

    # --- Rule 4: ZSP ---
    if student.zsp_completed_date is None and zsp_subject_name != "N/A":
        add_practicum_demand(
            "ZSP", program_type, zsp_subject_name, demand_counts, code_to_display_map
        )


def add_practicum_demand(
    practicum_type, program_type, subject, demand_counts, code_to_display_map
):
    """Adds a practicum demand entry and updates display mapping."""
    code = get_subject_code(program_type, practicum_type, subject)
    display = get_subject_display_name(program_type, practicum_type, subject)
    code_to_display_map[code] = display
    key = practicum_type, program_type, code
    demand_counts[key] += 1


def format_demand(demand_counts, code_to_display_map):
    """Formats the aggregated demand into a list of dicts."""
    formatted_demand = []
    for (ptype, pprog, pcode), count in demand_counts.items():
        formatted_demand.append(
            {
                "practicum_type": ptype,
                "program_type": pprog,
                "subject_code": pcode,
                "subject_display_name": code_to_display_map.get(pcode, pcode),
                "required_slots": count,
            }
        )
    return formatted_demand


def _parse_constraints_from_notes(mentor: PraktikumsLehrkraft) -> dict:
    """
    Parses the 'besonderheiten' field for hard constraints and returns a structured dict.
    """
    notes_text = (mentor.current_year_notes or "").lower()

    if not notes_text:
        return {
            "is_unavailable": False,
            "force_capacity": None,
            "allowed_types": None,
            "forbidden_combinations": set(),
        }

    # 1. Global Unavailability (Critical Check)
    if _is_mentor_unavailable(notes_text):
        return {
            "is_unavailable": True,
            "force_capacity": None,
            "allowed_types": None,
            "forbidden_combinations": set(),
        }

    # 2. specific constraints
    return {
        "is_unavailable": False,
        "force_capacity": _parse_capacity_override(notes_text),
        "allowed_types": _parse_allowed_types(notes_text),
        "forbidden_combinations": _parse_forbidden_combinations(notes_text),
    }


def _is_mentor_unavailable(text: str) -> bool:
    """Checks for keywords indicating the mentor cannot take students."""
    unavailable_keywords = ["ruhend", "sabbatjahr", "mobil", "elternzeit", "krank"]
    return any(keyword in text for keyword in unavailable_keywords)


def _parse_capacity_override(text: str):
    """Checks if the notes specify a strict capacity limit."""
    if "nur 1 prak" in text:
        return 1
    return None


def _parse_allowed_types(text: str):
    """Determines if the mentor is restricted to specific internship types."""
    if "nur blockpraktika" in text or "nur pdp" in text or "kein mi-prak" in text:
        return ["PDP_I", "PDP_II"]

    if "nur mi-prak" in text or "wg. tz nur mi-prak" in text:
        return ["SFP", "ZSP"]

    return None


def _parse_forbidden_combinations(text: str) -> set:
    """Identifies specific Subject+Type combinations that are banned."""
    forbidden = set()

    if "sfp nicht in geschichte" in text:
        forbidden.add(("SFP", "GE"))

    if (
        "englisch nicht möglich" in text
        or "englisch wird heuer nicht unterrichtet" in text
    ):
        forbidden.add(("SFP", "E"))
        forbidden.add(("ZSP", "E"))

    if "heuer kein krel" in text:
        forbidden.add(("SFP", "KRel"))
        forbidden.add(("ZSP", "KRel"))

    return forbidden


def calculate_eligibility_for_pl(mentor: PraktikumsLehrkraft) -> list:
    """
    Calculates the complete set of valid (Practicum Type Code, Subject Code)
    combinations for a given mentor, applying constraints from the 'besonderheiten' field.
    """
    # 1. Get all constraints from the 'besonderheiten' field
    constraints = _parse_constraints_from_notes(mentor)
    if not mentor.is_active or constraints.get("is_unavailable"):
        return []

    # 2. Determine base set of applicable internship types
    applicable_types = mentor.available_praktikum_types.all()
    if constraints.get("allowed_types"):
        applicable_types = applicable_types.filter(
            code__in=constraints["allowed_types"]
        )

    # 3. Generate all potential combinations
    potential_combinations = set()
    for p_type in applicable_types:
        reachable_schools = get_reachable_schools(p_type.code)
        # If the mentor's school is NOT in this list, they are not eligible for this type
        if mentor.school not in reachable_schools:
            continue

        if p_type.is_block_praktikum:
            potential_combinations.add((p_type.code, "N/A"))
        else:  # SFP/ZSP
            for subject in mentor.available_subjects.all():
                potential_combinations.add((p_type.code, subject.code))

    # 4. Filter out forbidden combinations found in the notes
    final_combinations = [
        combo
        for combo in potential_combinations
        if combo not in constraints["forbidden_combinations"]
    ]

    return final_combinations


def get_mentor_capacity(mentor: PraktikumsLehrkraft) -> int:
    """
    Gets the final capacity for a mentor, considering text-based overrides
    from the 'besonderheiten' field.
    """
    constraints = _parse_constraints_from_notes(mentor)
    if constraints.get("force_capacity") is not None:
        return constraints["force_capacity"]

    # Default to the capacity from Anrechnungsstunden
    return mentor.capacity


# ==================== DEMAND PREVIEW API ====================

def _calculate_available_pls_for_demand_item(demand_item):
    """
    Calculate the number of eligible PLs for a given demand item.
    Business Logic: PDP I/II only need program match, SFP/ZSP need subject match.
    
    Args:
        demand_item: dict with practicum_type, program_type, subject_code
    
    Returns:
        int: count of eligible PLs
    """
    eligible_pls_query = PraktikumsLehrkraft.objects.filter(
        is_active=True,
        program=demand_item['program_type']
    )
    
    # Add subject filter ONLY for SFP and ZSP
    if demand_item['practicum_type'] in ['SFP', 'ZSP']:
        eligible_pls_query = eligible_pls_query.filter(
            available_subjects__code=demand_item['subject_code']
        )
    
    return eligible_pls_query.distinct().count()


def _build_detailed_breakdown(raw_demand):
    """
    Build detailed breakdown by adding available_pls to each demand item.
    Business Logic: For each demand group, calculate supply from PLs.
    
    Args:
        raw_demand: list of demand dicts from aggregate_demand()
    
    Returns:
        list: demand items with available_pls added
    """
    detailed_breakdown = []
    
    for item in raw_demand:
        available_pls = _calculate_available_pls_for_demand_item(item)
        breakdown_item = {
            'practicum_type': item['practicum_type'],
            'program_type': item['program_type'],
            'subject_code': item['subject_code'],
            'subject_display_name': item['subject_display_name'],
            'required_slots': item['required_slots'],
            'available_pls': available_pls,
        }
        detailed_breakdown.append(breakdown_item)
    
    return detailed_breakdown


def _calculate_total_pl_capacity():
    """
    Calculate total capacity of all active PLs.
    Business Logic: Sum of (anrechnungsstunden * 2) for each active PL.
    
    Returns:
        int: total capacity slots
    """
    active_pls = PraktikumsLehrkraft.objects.filter(is_active=True)
    return sum(pl.capacity for pl in active_pls)


def _count_pls_for_block_praktikums():
    """
    Count PLs who can supervise block praktikums (PDP I, PDP II).
    Business Logic: Count active PLs with PDP_I or PDP_II in their available types.
    
    Returns:
        int: count of PLs available for block praktikums
    """
    return PraktikumsLehrkraft.objects.filter(
        is_active=True,
        available_praktikum_types__code__in=['PDP_I', 'PDP_II']
    ).distinct().count()


def _count_pls_for_wednesday_praktikums():
    """
    Count PLs who can supervise Wednesday praktikums (SFP, ZSP).
    Business Logic: Count active PLs with SFP or ZSP in their available types.
    
    Returns:
        int: count of PLs available for Wednesday praktikums
    """
    return PraktikumsLehrkraft.objects.filter(
        is_active=True,
        available_praktikum_types__code__in=['SFP', 'ZSP']
    ).distinct().count()


def _calculate_summary_cards(detailed_breakdown, total_pl_capacity):
    """
    Calculate summary card metrics from detailed breakdown.
    Business Logic: Aggregate demand counts and PL supply by practicum type categories.
    
    Args:
        detailed_breakdown: list of demand items with required_slots
        total_pl_capacity: pre-calculated total PL capacity
    
    Returns:
        dict: summary_cards with all metrics
    """
    total_demand_slots = 0
    
    for item in detailed_breakdown:
        total_demand_slots += item['required_slots']
    
    total_pdp_pls = _count_pls_for_block_praktikums()
    total_wednesday_pls = _count_pls_for_wednesday_praktikums()
    
    return {
        'total_demand_slots': total_demand_slots,
        'total_pl_capacity_slots': total_pl_capacity,
        'total_pdp_demand': total_pdp_pls,
        'total_wednesday_demand': total_wednesday_pls,
    }


def get_demand_preview_data():
    """
    Main service function to get complete demand preview data.
    Business Logic: Aggregates student demand and PL supply for allocation page.
    
    Returns:
        dict: {summary_cards: {...}, detailed_breakdown: [...]}
    """
    # Part 1: Get raw demand and build detailed breakdown
    raw_demand = aggregate_demand()
    detailed_breakdown = _build_detailed_breakdown(raw_demand)
    
    # Part 2: Calculate summary cards
    total_pl_capacity = _calculate_total_pl_capacity()
    summary_cards = _calculate_summary_cards(detailed_breakdown, total_pl_capacity)
    
    return {
        'summary_cards': summary_cards,
        'detailed_breakdown': detailed_breakdown,
    }


# ==================== EXPORT SERVICES ====================

def generate_assignments_excel():
    """
    Generate Excel export of all assignments.
    Business Logic: Export assignment master list with all details for archiving.
    
    Returns:
        bytes: Excel file content as bytes
    """
    from .models import Assignment
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    # Create workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Praktikumszuteilungen"
    
    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write header
    headers = ['ID', 'Praktikumstyp', 'Fach', 'Praktikumslehrkraft', 'Schule', 'Status']
    ws.append(headers)
    
    # Apply header styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Fetch assignments with related data
    assignments = Assignment.objects.select_related(
        'mentor', 'practicum_type', 'subject', 'school'
    ).all()
    
    # Write data rows
    for assignment in assignments:
        ws.append([
            assignment.id,
            assignment.practicum_type.get_code_display(),
            assignment.subject.code if assignment.subject else 'N/A',
            f"{assignment.mentor.last_name}, {assignment.mentor.first_name}",
            assignment.school.name,
            'Zugewiesen'
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def generate_assignments_pdf():
    """
    Generate PDF export of all assignments.
    Business Logic: Create formatted PDF reports for official documentation.
    
    Returns:
        bytes: PDF content as bytes
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.units import cm
    from io import BytesIO
    from .models import Assignment
    from datetime import datetime
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    # Add title
    title = Paragraph(
        f"<b>Praktikumszuteilungen - {datetime.now().strftime('%d.%m.%Y')}</b>",
        styles['Title']
    )
    elements.append(title)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Prepare table data
    data = [['ID', 'Typ', 'Fach', 'PL', 'Schule', 'Status']]
    
    assignments = Assignment.objects.select_related(
        'mentor', 'practicum_type', 'subject', 'school'
    ).all()
    
    for assignment in assignments:
        data.append([
            str(assignment.id),
            assignment.practicum_type.get_code_display(),
            assignment.subject.code if assignment.subject else 'N/A',
            f"{assignment.mentor.last_name}, {assignment.mentor.first_name}",
            assignment.school.name[:30],  # Truncate long names
            'OK'
        ])
    
    # Create and style table
    table = Table(data, colWidths=[1.5*cm, 3*cm, 2*cm, 5*cm, 6*cm, 2*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return buffer.getvalue()


def update_assignment(assignment_id, update_data):
    """
    Business Logic: Updates an assignment with new values.
    Allows editing mentor, school, subject, and practicum type.
    """
    from .models import Assignment
    from schools.models import School
    from subjects.models import Subject, PraktikumType

    try:
        assignment = Assignment.objects.get(id=assignment_id)
    except Assignment.DoesNotExist:
        return {"error": "Assignment not found"}, 404

    # Update mentor if provided
    if "mentor_id" in update_data:
        mentor = _get_mentor(update_data["mentor_id"])
        if not mentor:
            return {"error": "Mentor not found"}, 400
        assignment.mentor = mentor

    # Update school if provided
    if "school_id" in update_data:
        school = _get_school(update_data["school_id"])
        if not school:
            return {"error": "School not found"}, 400
        assignment.school = school

    # Update subject if provided
    if "subject_id" in update_data:
        subject = _get_subject(update_data["subject_id"])
        assignment.subject = subject

    # Update practicum type if provided
    if "practicum_type_id" in update_data:
        practicum_type = _get_practicum_type(update_data["practicum_type_id"])
        if not practicum_type:
            return {"error": "Practicum type not found"}, 400
        assignment.practicum_type = practicum_type

    assignment.save()
    return {"success": True, "assignment_id": assignment.id}, 200


def _get_mentor(mentor_id):
    """Helper: Fetches mentor by ID."""
    try:
        return PraktikumsLehrkraft.objects.get(id=mentor_id)
    except PraktikumsLehrkraft.DoesNotExist:
        return None


def _get_school(school_id):
    """Helper: Fetches school by ID."""
    from schools.models import School
    try:
        return School.objects.get(id=school_id)
    except School.DoesNotExist:
        return None


def _get_subject(subject_id):
    """Helper: Fetches subject by ID or returns None."""
    from subjects.models import Subject
    if subject_id is None:
        return None
    try:
        return Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return None


def _get_practicum_type(practicum_type_id):
    """Helper: Fetches practicum type by ID."""
    from subjects.models import PraktikumType
    try:
        return PraktikumType.objects.get(id=practicum_type_id)
    except PraktikumType.DoesNotExist:
        return None

